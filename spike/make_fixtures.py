"""Generate synthetic tax-style fixture PDFs for the Phase 0 spike.

NO client data — everything here is fabricated. Fixture design is deliberate:

fixture_a.pdf ("TaxForm-A") — 3 portrait pages, Rotate 0.
  Page sizes 612x792 / 613x792 / 614x792: unique MediaBox widths give every
  page a provenance signature we can assert after merge/reorder.

fixture_b.pdf ("SupportSchedules-B") — 3 pages, the hostile cases:
  B0: MediaBox 700x900 with CropBox [44,50,656,842]  -> CropBox != MediaBox
      (the coordinate-normalization trap named in the review).
  B1: 612x792 with /Rotate 90                        -> rotated-scan case.
  B2: 612x1008 (legal) carrying EXISTING annotations (Square + Text note)
      -> must survive merge/reorder (review pushback #2).
  Document outline: "Schedule X" -> B0 (child "Detail X-1" -> B1),
      "Schedule Y" -> B1 — imported-outline nesting/retargeting test.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pikepdf
from pikepdf import Array, Dictionary, Name, String

REPO = Path(__file__).resolve().parent.parent
FIXTURES = Path(__file__).parent / "fixtures"


def _esc(text: str) -> str:
    return text.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")


def _text_content(lines: list[tuple[float, float, float, str]]) -> bytes:
    """lines: (x, y, size, text) drawn in Helvetica."""
    parts = ["q 0.1 0.1 0.12 rg"]
    for x, y, size, text in lines:
        parts.append(f"BT /F1 {size:g} Tf {x:g} {y:g} Td ({_esc(text)}) Tj ET")
    parts.append("Q")
    return " ".join(parts).encode("ascii")


def _font_resources() -> Dictionary:
    return Dictionary(
        Font=Dictionary(
            F1=Dictionary(Type=Name.Font, Subtype=Name.Type1, BaseFont=Name.Helvetica)
        )
    )


def _add_page(
    pdf: pikepdf.Pdf,
    w: float,
    h: float,
    label: str,
    body: list[tuple[float, float, float, str]],
    rotate: int = 0,
    cropbox: list[float] | None = None,
) -> pikepdf.Page:
    page = pdf.add_blank_page(page_size=(w, h))
    lines = [(36.0, h - 60.0, 22.0, label)] + body
    page.obj.Contents = pdf.make_stream(_text_content(lines))
    page.obj.Resources = _font_resources()
    if rotate:
        page.obj.Rotate = rotate
    if cropbox:
        page.obj.CropBox = Array(cropbox)
    return page


def _tax_lines(x: float, y0: float, rows: list[tuple[str, str]]) -> list:
    out = []
    y = y0
    for label, amount in rows:
        out.append((x, y, 10.0, label))
        out.append((x + 340, y, 10.0, amount))
        y -= 16
    return out


def make_fixture_a(path: Path) -> None:
    with pikepdf.new() as pdf:
        _add_page(
            pdf, 612, 792, "TaxForm-A  p.A-1",
            _tax_lines(72, 690, [
                ("1  Wages, salaries, tips", "84,200.00"),
                ("2b Taxable interest", "1,150.00"),
                ("8  Other income (Sch 1)", "3,400.00"),
                ("11 Adjusted gross income", "88,750.00"),
            ]),
        )
        _add_page(
            pdf, 613, 792, "TaxForm-A  p.A-2",
            _tax_lines(72, 690, [
                ("12 Standard deduction", "14,600.00"),
                ("15 Taxable income", "74,150.00"),
                ("16 Tax", "11,807.00"),
            ]),
        )
        _add_page(
            pdf, 614, 792, "TaxForm-A  p.A-3",
            _tax_lines(72, 690, [
                ("25 Federal withholding", "12,400.00"),
                ("33 Total payments", "12,400.00"),
                ("34 Overpayment", "593.00"),
            ]),
        )
        pdf.save(path)


def make_fixture_b(path: Path) -> None:
    with pikepdf.new() as pdf:
        # B0 — CropBox != MediaBox. Content only in the TOP third of the crop
        # area so the spike's tape-region pixel check (lower area) stays clean.
        _add_page(
            pdf, 700, 900, "SupportSchedules-B  p.B-1",
            _tax_lines(80, 780, [
                ("Interest - First Natl", "612.00"),
                ("Interest - Credit Union", "538.00"),
                ("Total interest", "1,150.00"),
            ]),
            cropbox=[44, 50, 656, 842],
        )
        # B1 — rotated scan.
        _add_page(
            pdf, 612, 792, "SupportSchedules-B  p.B-2 (rotated scan)",
            _tax_lines(72, 640, [
                ("1099-MISC box 3", "3,400.00"),
                ("Agrees to Form line 8", ""),
            ]),
            rotate=90,
        )
        # B2 — carries pre-existing annotations.
        page3 = _add_page(
            pdf, 612, 1008, "SupportSchedules-B  p.B-3 (prior-year marked)",
            _tax_lines(72, 900, [
                ("Carryforward schedule", ""),
                ("Prior-year overpayment applied", "0.00"),
            ]),
        )
        square = pdf.make_indirect(Dictionary(
            Type=Name.Annot, Subtype=Name.Square,
            Rect=Array([100, 850, 220, 910]),
            C=Array([1, 0, 0]), F=4,
            NM=String("legacy-square-1"),
            Contents=String("Prior-year highlight"),
        ))
        note = pdf.make_indirect(Dictionary(
            Type=Name.Annot, Subtype=Name.Text,
            Rect=Array([240, 880, 262, 902]),
            Name=Name.Comment, F=4,
            NM=String("legacy-note-1"),
            Contents=String("Reviewed last year - agreed to bank stmt"),
        ))
        page3.obj.Annots = pdf.make_indirect(Array([square, note]))

        # B's own outline (imported-outline nesting test).
        from pikepdf import OutlineItem

        def fit(i: int) -> Array:
            return Array([pdf.pages[i].obj, Name.Fit])

        with pdf.open_outline() as outline:
            sched_x = OutlineItem("Schedule X", fit(0))
            sched_x.children.append(OutlineItem("Detail X-1", fit(1)))
            sched_y = OutlineItem("Schedule Y", fit(1))
            outline.root.append(sched_x)
            outline.root.append(sched_y)
        pdf.save(path)


def make_image_fixtures(landscape: Path, rotated: Path, png: Path) -> None:
    """Image sources — the receipt-photo and screenshot cases.

    Each carries a red block in its TOP-LEFT corner, which is what makes
    orientation assertable after export: if EXIF handling or the page /Rotate is
    wrong, the block lands on the wrong side of the sheet.

      receipt.jpg      400x300 landscape, no EXIF   -> landscape Letter page
      receipt_rot.jpg  same pixels, EXIF orient 6   -> portrait page via /Rotate,
                                                       embedded losslessly
      screenshot.png   200x500 portrait, has alpha  -> portrait page, re-encoded
                                                       (PNG isn't DCT) and
                                                       flattened onto white
    """
    from PIL import Image

    photo = Image.new("RGB", (400, 300), (255, 255, 255))
    for x in range(120):
        for y in range(80):
            photo.putpixel((x, y), (220, 30, 30))
    photo.save(landscape, quality=92)

    exif = photo.getexif()
    exif[274] = 6  # rotate 90 CW on display — a phone held portrait
    photo.save(rotated, quality=92, exif=exif)

    shot = Image.new("RGBA", (200, 500), (0, 0, 255, 255))
    for x in range(60):
        for y in range(60):
            shot.putpixel((x, y), (220, 30, 30, 255))
    shot.save(png)


def make_scan_fixture(source: Path, path: Path) -> None:
    """A page with NO text layer whose contents we nevertheless know exactly.

    Rasterizing a fixture we generated gives OCR a realistic scan while keeping
    ground truth: the figures are the ones make_fixture_a drew, so an OCR check
    can assert WHAT was read rather than merely that something was. Greyscale
    JPEG at 200 dpi is what a desk scanner actually produces.
    """
    import pypdfium2 as pdfium

    sys.path.insert(0, str(REPO / "engine"))
    from workpaper_engine import images

    doc = pdfium.PdfDocument(str(source))
    try:
        image = doc[0].render(scale=200 / 72).to_pil().convert("L")
    finally:
        doc.close()
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=88)
    jpg = path.with_name("scan_page.jpg")
    jpg.write_bytes(buf.getvalue())
    # Through the engine's own image path, so the result is a page exactly like
    # any photographed receipt a user would import.
    with images.image_to_pdf(str(jpg)) as pdf:
        pdf.save(path)


def make_workbook_fixtures(book: Path, register: Path) -> None:
    """A trial balance (short and wide) and a transaction register (long and
    narrow) — the two shapes a workpaper workbook actually comes in, and the
    reason page orientation is chosen per sheet rather than fixed.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    for row in [
        ["Account", "Description", "Debit", "Credit"],
        ["1000", "Cash - operating", 84200.00, None],
        ["1200", "Accounts receivable", 41850.25, None],
        ["1500", "Equipment", 128000.00, None],
        ["2000", "Accounts payable", None, 33110.40],
        ["3000", "Retained earnings", None, 220939.85],
    ]:
        ws.append(row)
    # Left uncalculated on purpose: openpyxl writes no cached value, which is
    # exactly the case that would otherwise render a blank where a total belongs.
    ws.append(["", "TOTAL", "=SUM(C2:C4)", "=SUM(D5:D6)"])
    dep = wb.create_sheet("Depreciation")
    dep.append(["Asset", "Cost", "Life", "Current"])
    dep.append(["Truck", 45000, 5, 9000])
    dep.append(["Server rack", 12500, 3, 4166.67])
    wb.save(book)

    wb2 = openpyxl.Workbook()
    reg = wb2.active
    reg.title = "Categorized Transactions"
    reg.append(["Date", "Description", "Source", "Acct #", "Account Name", "Debit", "Credit"])
    vendors = [
        "COMPUTER STORE", "ONLINE RETAILER - RETURN", "DOMAIN REGISTRAR", "PAYROLL SERVICE FEE",
        "CARD PAYMENT", "DESIGN SUBSCRIPTION", "SOCIAL PLATFORM - PAID", "DOCS PLATFORM, INC.",
    ]
    accounts = [
        (7260, "Small Equipment - De Minimis"), (7250, "Office Expenses"),
        (7300, "Software Subscriptions"), (7100, "Dues"), (7080, "Payroll Service Fees"),
    ]
    # 86 rows: enough that a fixed landscape page split it into a full page plus
    # a 13-row orphan, which is the bug that made fitting adaptive.
    for i in range(86):
        acct = accounts[i % len(accounts)]
        reg.append([
            f"2026-0{3 + i % 4}-{10 + i % 18:02d}",
            vendors[i % len(vendors)],
            "CARD" if i % 2 else "Operating #1010",
            acct[0],
            acct[1],
            round(25 + (i * 37.5) % 1500, 2),
            None,
        ])
    wb2.save(register)


def make_tb_columns_fixture(path: Path) -> None:
    """A trial balance shaped the way a real one is: a title block above the
    header, account numbers, and sparse debit/credit column PAIRS.

    This is the shape that showed the rendered page is not enough for an agent.
    Flattened to a line, "1001 Cash 7,412.68 5,310.40 4,982.15 7,740.93" loses
    which figure is a beginning balance and which is an ending one, because the
    blank cells disappear — and on a trial balance that is the whole meaning.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance 6-30-26"
    for row in [
        ["TRILAND PARTNERS LLC"],
        ["Trial Balance"],
        ["As of June 30, 2026"],
        [],
        ["Acct #", "Account Name", "Beg Dr", "Beg Cr", "Activity Dr", "Activity Cr",
         "Ending Dr", "Ending Cr"],
        [None, "BALANCE SHEET ACCOUNTS"],
        [1001, "Cash - Operating #1010", 7412.68, None, 5310.40, 4982.15, 7740.93, None],
        [1002, "Cash - Payroll #1020", 41.07, None, None, None, 41.07, None],
        [1500, "Accumulated Depreciation", None, 1734, None, None, None, 1734],
        [2000, "Credit Card Payable", None, 9.20, 4655.85, 5102.30, None, 455.65],
        [3000, "Retained Earnings", None, 5210.44, None, None, None, 5210.44],
    ]:
        ws.append(row)
    wb.save(path)


def make_memo_fixture(path: Path) -> None:
    """A memo of the shape an agent actually writes for an engagement.

    Exercises every block the renderer handles, because the failure mode is
    silent: an unhandled construct does not error, it just vanishes from the
    page, and a workpaper quietly missing a paragraph is worse than one that
    fails to import.
    """
    path.write_text(
        """# Q2 2026 Review Memo

Prepared to support the quarterly close. All figures agree to the general
ledger unless noted.

## Scope

- Reviewed **all** transactions over $500
- Reconciled the card and operating accounts through 6/30

## Findings

| Account | Description | Amount |
| --- | --- | --- |
| 7300 | Software Subscriptions | 1,203.26 |
| 7260 | Small Equipment | 2,808.76 |

### Software subscriptions

The balance is materially *higher* than Q1 because of the annual `ANTHROPIC`
renewal, which was expensed rather than prepaid.

See [the restatement memo](../../equity-restatement-2026.md) for the history,
and [IRC 1362](https://www.law.cornell.edu/uscode/text/26/1362) for authority.

> Management represents that no subscription exceeds twelve months.

## Conclusion

1. The trial balance foots and agrees to the detail.
2. No adjusting entries are proposed.

---

Prepared by ABC
""",
        encoding="utf-8",
    )


def make_multipage_tiff_fixture(path: Path) -> None:
    """A 3-frame TIFF, the standard output of an office scanner batch.

    Import must REFUSE it with an actionable message — the old behavior kept
    frame 0 and silently dropped the rest, which for a 40-page scan batch is
    39 missing client documents that nothing ever reports.
    """
    from PIL import Image, ImageDraw

    frames = []
    for i in range(3):
        img = Image.new("L", (400, 550), 255)
        ImageDraw.Draw(img).text((30, 30), f"Scan page {i + 1} of 3", fill=0)
        frames.append(img)
    frames[0].save(path, save_all=True, append_images=frames[1:])


def make_form_fixture(path: Path) -> None:
    """One page with a FILLED AcroForm text field relying on /NeedAppearances.

    The blocker this guards: `pages.append` copies the widget annotation but
    not the document-level /AcroForm, so the filled value renders as a blank
    box — a W-9 or 8879 silently loses the client's answers on export.
    NeedAppearances (no appearance stream of our own) is the hard variant:
    the value can ONLY render if /AcroForm and its /DR fonts survive.
    """
    with pikepdf.new() as pdf:
        pdf.add_blank_page(page_size=(612, 792))
        page = pdf.pages[0]
        field = pdf.make_indirect(
            Dictionary(
                FT=Name.Tx,
                T=String("business_name"),
                V=String("Whitmore Holdings LLC"),
                Type=Name.Annot,
                Subtype=Name.Widget,
                Rect=[100, 600, 500, 640],
                DA=String("/Helv 14 Tf 0 g"),
                F=4,
                P=page.obj,
            )
        )
        page.obj.Annots = pdf.make_indirect(Array([field]))
        helv = pdf.make_indirect(
            Dictionary(Type=Name.Font, Subtype=Name.Type1, BaseFont=Name.Helvetica)
        )
        pdf.Root.AcroForm = pdf.make_indirect(
            Dictionary(
                Fields=Array([field]),
                NeedAppearances=True,
                DA=String("/Helv 14 Tf 0 g"),
                DR=Dictionary(Font=Dictionary(Helv=helv)),
            )
        )
        pdf.save(path)


def main() -> dict[str, str]:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    a = FIXTURES / "fixture_a.pdf"
    b = FIXTURES / "fixture_b.pdf"
    make_fixture_a(a)
    make_fixture_b(b)
    img = FIXTURES / "receipt.jpg"
    img_rot = FIXTURES / "receipt_rot.jpg"
    shot = FIXTURES / "screenshot.png"
    make_image_fixtures(img, img_rot, shot)
    scan = FIXTURES / "scan_a.pdf"
    make_scan_fixture(a, scan)
    book = FIXTURES / "trial_balance.xlsx"
    register = FIXTURES / "long_register.xlsx"
    make_workbook_fixtures(book, register)
    columns = FIXTURES / "tb_columns.xlsx"
    make_tb_columns_fixture(columns)
    memo = FIXTURES / "review_memo.md"
    make_memo_fixture(memo)
    form = FIXTURES / "filled_form.pdf"
    make_form_fixture(form)
    scan_batch = FIXTURES / "scan_batch.tif"
    make_multipage_tiff_fixture(scan_batch)
    return {
        "A": str(a),
        "B": str(b),
        "IMG": str(img),
        "IMG_ROT": str(img_rot),
        "PNG": str(shot),
        "SCAN": str(scan),
        "BOOK": str(book),
        "REGISTER": str(register),
        "MEMO": str(memo),
        "TB_COLUMNS": str(columns),
        "FORM": str(form),
        "SCAN_BATCH": str(scan_batch),
    }


if __name__ == "__main__":
    print(main())
