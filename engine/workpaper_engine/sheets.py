"""Spreadsheets as binder pages.

Excel IS the workpaper format for most of what a preparer assembles — trial
balances, lead sheets, depreciation schedules. Until now none of it could enter
a binder at all, which made "bring in the documents I actually have" false.

Follows the same invariant as images: the session keeps pointing at the
untouched .xlsx, and the pages are built in memory at export. Nothing here
rewrites a source file.

TWO DELIBERATE CHOICES:

1. **We lay the grid out ourselves rather than shelling out to LibreOffice or
   Excel.** A 500 MB office suite cannot be bundled, and requiring an install
   repeats the problem OCR already has. The honest cost: this is a clean
   rendering of the DATA, not Excel's own print layout — no merged-cell art, no
   conditional formatting, no charts. For a trial balance that is fine and
   arguably more legible; for a formatted client-facing schedule it is not, and
   the user should print that one to PDF themselves.

2. **Courier, not Helvetica.** Standard-14 so nothing is embedded, and every
   glyph is exactly 0.6 em — so column fitting is exact arithmetic instead of a
   font-metrics table, and figures line up in columns the way an accountant
   reads them.

Because the text is really drawn into the page, the existing text extraction
finds it with exact positions: a figure from a spreadsheet is addressable and
tickable with no OCR involved.
"""

from __future__ import annotations

import codecs
import csv as csvlib
import datetime as dt
import io
from pathlib import Path

import pikepdf
from pikepdf import Array, Dictionary, Name

from .appearance import _esc

SHEET_SUFFIXES = frozenset({".xlsx", ".xlsm", ".csv"})

# Both Letter orientations are tried per sheet: a transaction register is tall
# and narrow, a trial balance short and wide, and forcing one shape on both is
# what turned an 87-row sheet into a full page plus a 13-row orphan.
LANDSCAPE = (792.0, 612.0)
PORTRAIT = (612.0, 792.0)
MARGIN = 24.0
FONT_MAX = 9.0
FONT_MIN = 5.5
# How small we will go to keep a sheet on ONE page. Below this it stops being
# something a preparer can read, and two honest pages beat one unreadable one.
ONE_PAGE_MIN = 4.5
LINE_GAP = 1.35
CHAR_W = 0.6  # Courier advance, in em — exact, not an approximation
GRID_GREY = 0.75
# A runaway sheet must not turn one import into a thousand pages.
MAX_PAGES_PER_SHEET = 200
MAX_ROWS = 20000
MAX_COLS = 256


def is_sheet(path: str | Path) -> bool:
    return Path(path).suffix.lower() in SHEET_SUFFIXES


def _clean(value, number_format: str = "") -> str:
    """One cell as a string a Courier/WinAnsi page can actually show.

    `number_format` is the workbook's own format for the cell, and it decides
    whether a whole number is grouped. Guessing gets account numbers wrong:
    1001 rendered as "1,001" is not a formatting nicety, it is a value an agent
    searching for account 1001 will never find.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    grouped = "#,#" in (number_format or "")
    if isinstance(value, float):
        # Trailing-zero noise ("1234.5600000000001") is not what was in the cell.
        text = (
            f"{value:,.2f}"
            if value != int(value)
            else (f"{int(value):,}" if grouped else f"{int(value)}")
        )
    elif isinstance(value, int):
        text = f"{value:,}" if grouped else f"{value}"
    elif isinstance(value, (dt.datetime, dt.date)):
        text = value.strftime("%Y-%m-%d")
    else:
        text = str(value)
    # This used to end `ch if 32 <= ord(ch) < 256 else "?"`, reasoning that
    # WinAnsi cannot show every codepoint and a visible marker beats a broken
    # glyph. The reasoning was right; the TEST was wrong. A codepoint's numeric
    # value does not say whether WinAnsi carries it: the em dash is U+2014, so
    # it failed `< 256` and became "?" — while WinAnsi carries it perfectly well
    # at 0x97. "Interest — Sch B" in a trial balance rendered "Interest ? Sch B"
    # and nothing said so. Meanwhile codepoints like U+0081 passed the test and
    # WinAnsi genuinely cannot carry them.
    #
    # `_esc` answers the real question — it asks cp1252 whether it can encode
    # the character, and writes a visible [U+XXXX] when it cannot. That is the
    # same "visible marker" this line wanted, applied on the right test, so the
    # substitution belongs there and not here. Control characters still go, as
    # they would break the content stream rather than merely look wrong.
    return "".join(ch for ch in text.replace("\n", " ") if ord(ch) >= 32 or ch == "\t")


def _decode_csv(raw: bytes) -> tuple[str, str | None]:
    """Decode CSV bytes, returning the text and a warning if it had to guess.

    This used to be `encoding="utf-8-sig", errors="replace"` — one encoding
    assumed, and anything that was not it replaced character by character. The
    assumption is the wrong way round for this profession. Excel's plain "CSV"
    export writes the WINDOWS ANSI codepage, not UTF-8, and it is the export a
    firm reaches for; only the separately-named "CSV UTF-8" writes UTF-8. So a
    trial balance for Peña & Fuentes, saved the ordinary way, arrived as
    "Pe�a" — every accented character replaced before the file was even
    parsed, with no error and nothing to notice.

    The order below is what the bytes can actually support, not a preference:

    1. A BOM is a statement of fact; believe it.
    2. UTF-8, strictly. This is self-validating — its multi-byte sequences are
       structured enough that ordinary cp1252 text almost never decodes as
       valid UTF-8 by accident — so a clean decode is strong evidence, and a
       failure is proof it is NOT UTF-8 rather than a reason to start replacing
       characters.
    3. cp1252, strictly. What Excel actually wrote.
    4. latin-1, which cannot fail, so there is always an answer.

    Steps 3 and 4 are guesses and say so. Encoding cannot be recovered from
    bytes with certainty; what it can do is stop silently destroying them.
    """
    for bom, enc in (
        (codecs.BOM_UTF8, "utf-8-sig"),
        (codecs.BOM_UTF16_LE, "utf-16"),
        (codecs.BOM_UTF16_BE, "utf-16"),
    ):
        if raw.startswith(bom):
            return raw.decode(enc), None
    try:
        return raw.decode("utf-8"), None
    except UnicodeDecodeError:
        pass
    try:
        # Kept short on purpose: this lands in the status bar, which truncates.
        # A caveat the preparer cannot finish reading is half a caveat.
        return raw.decode("cp1252"), "read as Windows-1252, not UTF-8 — check accented names"
    except UnicodeDecodeError:
        return raw.decode("latin-1"), "read as Latin-1 — check any non-English characters"


def read_grids(path: str | Path) -> tuple[list[tuple[str, list[list[str]]]], list[str]]:
    """(sheet name, rows of cell strings) per sheet, plus any warnings."""
    p = Path(path)
    warnings: list[str] = []
    if p.suffix.lower() == ".csv":
        text, guess = _decode_csv(p.read_bytes())
        if guess:
            warnings.append(f"{p.name}: {guess}")
        rows = [
            [_clean(cell) for cell in row[:MAX_COLS]]
            for row in list(csvlib.reader(io.StringIO(text, newline="")))[:MAX_ROWS]
        ]
        return [(p.stem, rows)], warnings

    import openpyxl

    # data_only gives the value Excel last CALCULATED. A workbook written by a
    # tool that never calculated has none — the cell reads empty, which in a
    # workpaper is a blank where a number belongs. So the formulas are read too
    # and shown when their cached value is missing, flagged rather than hidden.
    values = openpyxl.load_workbook(p, data_only=True, read_only=True)
    formulas = openpyxl.load_workbook(p, data_only=False, read_only=True)
    grids: list[tuple[str, list[list[str]]]] = []
    uncalculated = 0
    try:
        for name in values.sheetnames:
            vsheet = values[name]
            fsheet = formulas[name] if name in formulas.sheetnames else None
            rows: list[list[str]] = []
            frows = fsheet.iter_rows(values_only=True) if fsheet else iter(())
            for row, frow in zip(vsheet.iter_rows(), frows):
                out: list[str] = []
                for i, cell in enumerate(row[:MAX_COLS]):
                    text = _clean(cell.value, getattr(cell, "number_format", ""))
                    if not text and i < len(frow):
                        raw = frow[i]
                        if isinstance(raw, str) and raw.startswith("="):
                            text = _clean(raw)
                            uncalculated += 1
                    out.append(text)
                rows.append(out)
                if len(rows) >= MAX_ROWS:
                    warnings.append(f"{name}: stopped at {MAX_ROWS} rows")
                    break
            grids.append((name, rows))
    finally:
        values.close()
        formulas.close()
    if uncalculated:
        warnings.append(
            f"{uncalculated} formula cell(s) had no calculated value and are shown "
            "as formulas — open and save the workbook in Excel to resolve them"
        )
    return grids, warnings


def _header_row(rows: list[list[str]]) -> int:
    """Which row names the columns.

    A workpaper sheet usually opens with a firm name, a title and a date before
    the real header. The header is taken to be the earliest row with at least
    two non-empty cells and no numbers in them — a title row has one cell, a
    data row has figures.
    """
    for i, row in enumerate(rows[:12]):
        filled = [c for c in row if c.strip()]
        if len(filled) < 2:
            continue
        numeric = sum(
            1
            for c in filled
            if c.replace(",", "").replace(".", "").replace("-", "").isdigit()
        )
        if numeric == 0:
            return i
    return 0


def read_cells(path: str | Path, max_rows: int = 2000) -> dict:
    """The sheet as DATA, not as a picture of a page.

    The rendered page flattens a row to "1001 Cash - Operating #1010 7,412.68
    5,310.40 4,982.15 7,740.93" — blank cells vanish, so nothing says which
    figure is Beg Dr and which is Ending Dr. On a trial balance that is the
    entire meaning, and an agent asked to reconcile has to guess.

    We parsed the real cells to build the page, so this hands them over rather
    than making an agent recover structure from a rendering. Empty cells are
    returned EXPLICITLY, because "this column is blank for this account" is a
    fact a reconciliation depends on.
    """
    grids, warnings = read_grids(path)
    out_sheets = []
    for name, raw in grids:
        rows = _trim(raw)
        if not rows:
            out_sheets.append({"name": name, "headers": [], "rows": [], "header_row": None})
            continue
        h = _header_row(rows)
        headers = [c.strip() or f"col{i + 1}" for i, c in enumerate(rows[h])]
        body = []
        for r, row in enumerate(rows[h + 1 :][:max_rows], start=h + 2):
            if not any(c.strip() for c in row):
                continue
            width = max(len(headers), len(row))
            body.append(
                {
                    # 1-based, matching what Excel shows in the row gutter.
                    "row": r,
                    "cells": {
                        (headers[i] if i < len(headers) else f"col{i + 1}"): (
                            row[i] if i < len(row) else ""
                        )
                        for i in range(width)
                    },
                }
            )
        out_sheets.append(
            {
                "name": name,
                # Reported so a caller can see what was guessed and disagree.
                "header_row": h + 1,
                "headers": headers,
                "rows": body,
                "truncated": len(rows) - h - 1 > max_rows,
            }
        )
    return {"sheets": out_sheets, "warnings": warnings}


def _trim(rows: list[list[str]]) -> list[list[str]]:
    """Drop the empty right and bottom margins openpyxl reports."""
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    width = max((max((i + 1 for i, c in enumerate(r) if c.strip()), default=0) for r in rows), default=0)
    return [r[:width] + [""] * (width - len(r)) for r in rows]


def _fits(rows, widest, page, size) -> tuple[bool, bool]:
    """(columns fit the width, all rows fit one page) at this size."""
    wide = sum((w + 2) * CHAR_W * size for w in widest) <= page[0] - 2 * MARGIN
    line = size * LINE_GAP
    tall = len(rows) * line <= page[1] - 2 * MARGIN - line
    return wide, tall


def _plan(rows: list[list[str]]) -> tuple[tuple[float, float], float, list[float], int, list[list[int]]]:
    """Choose page shape, font size, column widths, rows per page, column bands.

    ONE PAGE PER SHEET IS THE GOAL. A workbook that spills 13 rows onto a second
    page has not been paginated, it has been broken: the preparer gets a full
    page and an orphan. So the largest size that fits everything on a single
    page is found first, in BOTH Letter orientations — a transaction register is
    tall and narrow, a trial balance short and wide, and forcing one shape on
    both is what caused the orphan.

    Only when that would require type smaller than ONE_PAGE_MIN does it fall
    back to paginating, because two readable pages beat one that nobody can
    read.
    """
    if not rows:
        return LANDSCAPE, FONT_MAX, [], 1, [[]]
    n_cols = len(rows[0])
    # +2 chars of gutter, not +1: at 9pt one character is ~5pt and a label
    # ran straight into the figure beside it, which on a trial balance reads
    # as a single value.
    widest = [max((len(r[c]) for r in rows), default=0) for c in range(n_cols)]

    best: tuple[tuple[float, float], float] | None = None
    for page in (LANDSCAPE, PORTRAIT):
        size = FONT_MAX
        while size >= ONE_PAGE_MIN:
            wide, tall = _fits(rows, widest, page, size)
            if wide and tall:
                if best is None or size > best[1]:
                    best = (page, size)
                break
            size -= 0.25

    if best is not None:
        page, size = best
        cols = [(w + 2) * CHAR_W * size for w in widest]
        return page, size, cols, len(rows), [list(range(n_cols))]

    # Too much to fit legibly. Fall back to filling pages: pick the orientation
    # that needs fewer column bands, then paginate rows.
    page = LANDSCAPE
    size = FONT_MAX
    while size > FONT_MIN:
        if _fits(rows, widest, page, size)[0]:
            break
        size -= 0.5

    cols = [(w + 2) * CHAR_W * size for w in widest]
    usable = page[0] - 2 * MARGIN
    bands: list[list[int]] = []
    current: list[int] = []
    used = 0.0
    for i, width in enumerate(cols):
        # A single column wider than the page still gets its own band; it will
        # overflow rather than vanish.
        if current and used + width > usable:
            bands.append(current)
            current, used = [], 0.0
        current.append(i)
        used += width
    bands.append(current)

    line = size * LINE_GAP
    per_page = max(1, int((page[1] - 2 * MARGIN - line) // line))
    return page, size, cols, per_page, bands


def _numeric_columns(rows: list[list[str]], n_cols: int) -> list[bool]:
    """Which columns hold figures.

    Accountants read a column of numbers down its right edge — left-aligned
    money is why a rendered trial balance looks scattered. Decided per column
    from the data rather than per cell, so one stray note does not unalign the
    whole column.
    """
    out: list[bool] = []
    for c in range(n_cols):
        values = [r[c].strip() for r in rows if c < len(r) and r[c].strip()]
        if not values:
            out.append(False)
            continue
        numeric = sum(
            1
            for v in values
            if v.replace(",", "").replace(".", "").replace("-", "").replace("(", "").replace(")", "").isdigit()
        )
        out.append(numeric >= max(1, len(values) * 0.6))
    return out


#: The fourth copy of `_esc` lived here. b7763a1 consolidated the other three
#: into appearance.py and missed this one, so a spreadsheet kept its own rules:
#: delimiters escaped, everything else passed through raw and then encoded
#: `latin-1, "replace"` below — which is not a crash but is worse in one
#: specific way. An em dash is not IN latin-1, so "Interest — Sch B" in a cell
#: rendered as "Interest ? Sch B", and a trial balance for Peña & Fuentes came
#: out "Pe?a". A figure read off that page is still right; the label naming what
#: it is quietly is not, and nothing anywhere reported a problem.
#:
#: The shared one emits WinAnsi octal escapes, and this module's font already
#: declares /WinAnsiEncoding, so it needed no other change. A character WinAnsi
#: genuinely cannot carry becomes a visible [U+XXXX] — the honest failure the
#: rest of the engine uses, rather than a "?" indistinguishable from a "?" the
#: client actually typed.


def _page_stream(
    rows: list[list[str]],
    band: list[int],
    cols: list[float],
    size: float,
    title: str,
    page: tuple[float, float],
) -> bytes:
    parts: list[str] = []
    line = size * LINE_GAP
    y = page[1] - MARGIN - size

    parts.append(f"BT /F1 {size:g} Tf {MARGIN:g} {y:g} Td ({_esc(title)}) Tj ET")
    y -= line

    # A hairline under the header keeps a long table readable without pretending
    # to reproduce the workbook's own formatting.
    parts.append(f"q {GRID_GREY:g} G 0.4 w {MARGIN:g} {y + size * 0.9:g} m "
                 f"{page[0] - MARGIN:g} {y + size * 0.9:g} l S Q")

    right = _numeric_columns(rows, len(cols)) if cols else []
    for row in rows:
        x = MARGIN
        for c in band:
            text = row[c] if c < len(row) else ""
            if text:
                at = x
                if c < len(right) and right[c]:
                    # Right edge of the column, less one character of gutter.
                    at = x + cols[c] - (len(text) + 1) * CHAR_W * size
                parts.append(f"BT /F1 {size:g} Tf {at:g} {y:g} Td ({_esc(text)}) Tj ET")
            x += cols[c]
        y -= line
    # ASCII, not `latin-1, "replace"`. The shared `_esc` has already turned
    # everything outside ASCII into an octal escape, so there is nothing left
    # for a lossy codec to stand in for — and "replace" was the mechanism that
    # turned an em dash into "?" without telling anyone.
    return " ".join(parts).encode("ascii")


def sheet_to_pdf(path: str | Path) -> pikepdf.Pdf:
    """Render a spreadsheet to an in-memory PDF. The file on disk is untouched."""
    grids, _warnings = read_grids(path)
    pdf = pikepdf.new()
    font = pdf.make_indirect(
        Dictionary(
            Type=Name.Font,
            Subtype=Name.Type1,
            BaseFont=Name("/Courier"),
            Encoding=Name("/WinAnsiEncoding"),
        )
    )
    resources = Dictionary(Font=Dictionary(F1=font))

    for name, raw in grids:
        rows = _trim(raw)
        shape, size, cols, per_page, bands = _plan(rows)
        chunks = [rows[i : i + per_page] for i in range(0, len(rows), per_page)] or [[]]
        made = 0
        for band_i, band in enumerate(bands):
            for chunk_i, chunk in enumerate(chunks):
                if made >= MAX_PAGES_PER_SHEET:
                    break
                label = name
                if len(chunks) > 1:
                    label += f"  (rows {chunk_i * per_page + 1}-{chunk_i * per_page + len(chunk)})"
                if len(bands) > 1:
                    label += f"  (columns {band_i + 1} of {len(bands)})"
                sheet_page = pdf.add_blank_page(page_size=shape)
                sheet_page.obj.Contents = pdf.make_stream(
                    _page_stream(chunk, band, cols, size, label, shape)
                )
                sheet_page.obj.Resources = resources
                made += 1
    if len(pdf.pages) == 0:
        blank = pdf.add_blank_page(page_size=LANDSCAPE)
        blank.obj.Contents = pdf.make_stream(
            _page_stream([], [], [], FONT_MAX, Path(path).name, LANDSCAPE)
        )
        blank.obj.Resources = resources
    return pdf


def probe_sheet(path: str | Path) -> dict:
    """Same shape probe_image returns, so import treats a sheet like any source."""
    grids, warnings = read_grids(path)
    boxes: list[list[float]] = []
    with sheet_to_pdf(path) as pdf:
        n_pages = len(pdf.pages)
        for page in pdf.pages:
            boxes.append([float(v) for v in page.obj.MediaBox])
    return {
        "path": str(path),
        "n_pages": n_pages,
        "kind": "sheet",
        "pages": [
            {"index": i, "rotate": 0, "mediabox": boxes[i], "cropbox": None}
            for i in range(n_pages)
        ],
        "outline": [],
        "sheet": {
            "sheets": [
                {"name": name, "rows": len(_trim(rows)), "columns": len(_trim(rows)[0]) if _trim(rows) else 0}
                for name, rows in grids
            ],
            "warnings": warnings,
        },
    }
